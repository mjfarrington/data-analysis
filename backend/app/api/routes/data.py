from __future__ import annotations
import asyncio
import logging
import traceback
import csv
import io
import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse, FileResponse

from app.core.config import settings
from app.schemas.etl import QueryRequest, QueryResult, DataTable, BulkQueryRequest, QueryExportRequest
from app.services.spark_service import spark_service, reset_spark_session, _get_spark

router = APIRouter(prefix="/data", tags=["Data"])
logger = logging.getLogger(__name__)

_EXPORT_JOBS: dict[str, dict] = {}

_ALLOWED_SQL_PREFIXES = (
    "SELECT", "SHOW", "DESCRIBE", "EXPLAIN",
    "DROP", "CREATE", "ALTER", "INSERT", "UPDATE", "DELETE",
    "TRUNCATE", "WITH", "CACHE", "UNCACHE", "REFRESH", "ANALYZE",
    "USE", "SET", "RESET",
)


def _check_sql_allowed(sql: str) -> None:
    """Raise 400 if the SQL doesn't start with a recognised statement keyword."""
    stripped = sql.strip().upper()
    if not any(stripped.startswith(kw) for kw in _ALLOWED_SQL_PREFIXES):
        raise HTTPException(
            status_code=400,
            detail=f"Statement type not permitted. Allowed prefixes: {', '.join(_ALLOWED_SQL_PREFIXES)}",
        )


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _serialize_export_job(job: dict) -> dict:
    return {
        "id": job["id"],
        "status": job["status"],
        "phase": job.get("phase"),
        "format": job.get("format", "csv"),
        "rows_exported": job.get("rows_exported", 0),
        "max_rows": job.get("max_rows", 0),
        "database": job.get("database"),
        "created_at": job.get("created_at"),
        "started_at": job.get("started_at"),
        "completed_at": job.get("completed_at"),
        "updated_at": job.get("updated_at"),
        "error": job.get("error"),
        "download_ready": bool(job.get("file_path")) and job.get("status") == "completed",
    }


async def _run_query_export_job(job_id: str):
    job = _EXPORT_JOBS.get(job_id)
    if not job:
        return

    job["status"] = "running"
    job["phase"] = "querying"
    job["started_at"] = _utc_now_iso()
    job["updated_at"] = _utc_now_iso()

    try:
        result = await spark_service.execute_query_bulk(
            job["sql"],
            job["max_rows"],
            database=job.get("database"),
        )

        if job.get("cancel_requested"):
            job["status"] = "canceled"
            job["phase"] = "canceled"
            job["completed_at"] = _utc_now_iso()
            job["updated_at"] = _utc_now_iso()
            return

        columns = result.get("columns", [])
        rows = result.get("rows", [])
        job["rows_exported"] = len(rows)
        job["phase"] = "writing"
        job["updated_at"] = _utc_now_iso()

        export_dir = settings.data_path / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)

        fmt = job.get("format", "csv")

        if fmt == "xlsx":
            out_path = export_dir / f"query_export_{job_id}.xlsx"

            def _write_xlsx(path: Path, cols: list[str], data_rows: list[list]):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Export"

                # Palette
                HDR_BG = "1E3A5F"
                HDR_FG = "F8FAFC"
                BAND_BG = "EAF2FB"
                BORDER_COLOR = "C9D7E6"

                hdr_fill = PatternFill("solid", fgColor=HDR_BG)
                hdr_font = Font(bold=True, color=HDR_FG)
                band_fill = PatternFill("solid", fgColor=BAND_BG)
                thin = Side(style="thin", color=BORDER_COLOR)
                cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # Write header
                for col_idx, col_name in enumerate(cols, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                    cell.border = cell_border
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                # Write data rows
                for row_idx, row in enumerate(data_rows, start=2):
                    fill = band_fill if row_idx % 2 == 0 else None
                    for col_idx, value in enumerate(row, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        if fill:
                            cell.fill = fill
                        cell.border = cell_border
                        # Number formatting
                        if isinstance(value, float):
                            cell.number_format = '#,##0.00'
                        elif isinstance(value, int):
                            cell.number_format = '#,##0'

                # Auto-fit column widths (capped at 60)
                for col_idx, col_name in enumerate(cols, start=1):
                    col_letter = get_column_letter(col_idx)
                    max_len = len(str(col_name))
                    for row in data_rows:
                        val = row[col_idx - 1]
                        if val is not None:
                            max_len = max(max_len, len(str(val)))
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

                # Freeze header row and add autofilter
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

                wb.save(path)

            await asyncio.to_thread(_write_xlsx, out_path, columns, rows)
        else:
            out_path = export_dir / f"query_export_{job_id}.csv"

            def _write_csv(path: Path, cols: list[str], data_rows: list[list]):
                with path.open("w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(cols)
                    writer.writerows(data_rows)

            await asyncio.to_thread(_write_csv, out_path, columns, rows)

        if job.get("cancel_requested"):
            try:
                out_path.unlink(missing_ok=True)
            except Exception:
                pass
            job["status"] = "canceled"
            job["phase"] = "canceled"
            job["completed_at"] = _utc_now_iso()
            job["updated_at"] = _utc_now_iso()
            return

        job["status"] = "completed"
        job["phase"] = "completed"
        job["file_path"] = str(out_path)
        job["completed_at"] = _utc_now_iso()
        job["updated_at"] = _utc_now_iso()
    except Exception as exc:
        job["status"] = "failed"
        job["phase"] = "failed"
        job["error"] = str(exc)
        job["completed_at"] = _utc_now_iso()
        job["updated_at"] = _utc_now_iso()


def _suppress_jvm_stacktrace(text: str) -> str:
    """Remove noisy JVM stack-frame lines while keeping the root Spark error message."""
    if not text:
        return text
    cleaned: list[str] = []
    suppressed = False
    inserted_marker = False
    for line in text.splitlines():
        stripped = line.strip()
        is_jvm_frame = (
            stripped.startswith("at org.")
            or stripped.startswith("at java.")
            or stripped.startswith("at scala.")
            or stripped.startswith("at sun.")
            or (stripped.startswith("...") and stripped.endswith("more"))
        )
        if is_jvm_frame:
            suppressed = True
            if not inserted_marker:
                cleaned.append("[JVM stack trace suppressed]")
                inserted_marker = True
            continue
        cleaned.append(line)
    if not suppressed:
        return text
    return "\n".join(cleaned).strip()


# ─────────────────────────────────────────────────────────────────────────────
# Data Browser — list extracted parquet directories
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/browser")
async def list_browser_dirs():
    """List all parquet directories under PARQUET_DIR with basic metadata."""
    base = Path(settings.PARQUET_DIR)

    def _scan():
        entries = []
        if not base.exists():
            return entries
        for item in sorted(base.iterdir()):
            if not item.is_dir():
                continue
            parquet_files = sorted(item.glob("**/*.parquet"))
            csv_files = sorted(item.glob("**/*.csv")) if not parquet_files else []
            files = parquet_files or csv_files
            if not files:
                continue
            total_size = sum(f.stat().st_size for f in files)
            mtime = max(f.stat().st_mtime for f in files)
            entries.append({
                "name": item.name,
                "path": item.name,
                "file_count": len(files),
                "format": "parquet" if parquet_files else "csv",
                "size_bytes": total_size,
                "last_modified": mtime,
            })
        return entries

    return await asyncio.to_thread(_scan)


@router.get("/tables", response_model=list[DataTable])
async def list_data_tables():
    tables = await spark_service.list_tables()
    return [
        DataTable(
            name=t["name"],
            path=t["path"],
            format=t["format"],
            size_bytes=t["size_bytes"],
            row_count=t.get("row_count"),
            columns=t.get("columns", []),
            partitions=t.get("partitions", []),
            last_modified=t.get("last_modified"),
        )
        for t in tables
    ]


@router.get("/tables/{table_name:path}/preview")
async def preview_file_table(
    table_name: str,
    limit: int = Query(default=200, ge=1, le=2000),
    offset: int = Query(default=0, ge=0),
):
    """Preview a parquet/CSV file directly via pandas — no Spark required."""
    import pandas as pd

    # Resolve table_name (date/job/app_id) to the actual directory
    target = settings.parquet_path / table_name
    if not target.exists() or not target.is_dir():
        raise HTTPException(status_code=404, detail=f"No data directory for {table_name!r}")

    # Traverse path guard
    try:
        target.resolve().relative_to(settings.parquet_path.resolve())
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid table path")

    def _read() -> dict:
        parquet_files = sorted(target.glob("*.parquet"))
        csv_files = sorted(target.glob("*.csv"))

        if parquet_files:
            frames = [pd.read_parquet(f) for f in parquet_files]
            df = pd.concat(frames, ignore_index=True) if len(frames) > 1 else frames[0]
        elif csv_files:
            frames = [pd.read_csv(f) for f in csv_files]
            df = pd.concat(frames, ignore_index=True) if len(frames) > 1 else frames[0]
        else:
            raise FileNotFoundError(f"No parquet or CSV files in {target}")

        total_rows = len(df)
        page = df.iloc[offset: offset + limit]

        columns = list(page.columns)
        rows = []
        for _, row in page.iterrows():
            rows.append([None if pd.isna(v) else v for v in row.tolist()])

        return {
            "columns": columns,
            "rows": rows,
            "row_count": len(rows),
            "total_rows": total_rows,
            "truncated": (offset + len(rows)) < total_rows,
            "file_count": len(parquet_files) or len(csv_files),
            "format": "parquet" if parquet_files else "csv",
        }

    try:
        return await asyncio.to_thread(_read)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        logger.exception("Error previewing %s", table_name)
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/query", response_model=QueryResult)
async def execute_query(body: QueryRequest):
    _check_sql_allowed(body.sql)
    try:
        result = await spark_service.execute_query(body.sql, body.limit, body.offset, database=body.database)
        return QueryResult(**result)
    except Exception as exc:
        exc_name = exc.__class__.__name__
        spark_message = str(exc)
        spark_message_clean = _suppress_jvm_stacktrace(spark_message)
        msg_lower = spark_message.lower()
        is_user_sql_error = any(
            token in exc_name.lower() or token in msg_lower
            for token in ("parse", "analysis", "syntax", "sqlstate", "mismatched input")
        )
        traceback_clean = _suppress_jvm_stacktrace(traceback.format_exc())
        raise HTTPException(
            status_code=400 if is_user_sql_error else 500,
            detail={
                "message": "Spark SQL execution failed",
                "error_type": exc_name,
                "spark_message": spark_message_clean,
                "database": body.database,
                "sql": body.sql,
                "traceback": traceback_clean,
            },
        )


@router.post("/query/bulk", response_model=QueryResult)
async def execute_query_bulk(body: BulkQueryRequest):
    _check_sql_allowed(body.sql)
    try:
        result = await spark_service.execute_query_bulk(body.sql, body.max_rows, database=body.database)
        return QueryResult(**result)
    except Exception as exc:
        exc_name = exc.__class__.__name__
        spark_message = str(exc)
        spark_message_clean = _suppress_jvm_stacktrace(spark_message)
        msg_lower = spark_message.lower()
        is_user_sql_error = any(
            token in exc_name.lower() or token in msg_lower
            for token in ("parse", "analysis", "syntax", "sqlstate", "mismatched input")
        )
        traceback_clean = _suppress_jvm_stacktrace(traceback.format_exc())
        raise HTTPException(
            status_code=400 if is_user_sql_error else 500,
            detail={
                "message": "Spark SQL execution failed",
                "error_type": exc_name,
                "spark_message": spark_message_clean,
                "database": body.database,
                "sql": body.sql,
                "traceback": traceback_clean,
            },
        )


@router.post("/query/export")
async def export_query_file(body: QueryExportRequest):
    _check_sql_allowed(body.sql)
    try:
        result = await spark_service.execute_query_bulk(body.sql, body.max_rows, database=body.database)
        columns = result.get("columns", [])
        rows = result.get("rows", [])

        stream = io.StringIO()
        writer = csv.writer(stream)
        writer.writerow(columns)
        writer.writerows(rows)
        payload = io.BytesIO(stream.getvalue().encode("utf-8"))
        payload.seek(0)
        return StreamingResponse(
            payload,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=query_server_export.csv"},
        )
    except Exception as exc:
        exc_name = exc.__class__.__name__
        spark_message = str(exc)
        spark_message_clean = _suppress_jvm_stacktrace(spark_message)
        msg_lower = spark_message.lower()
        is_user_sql_error = any(
            token in exc_name.lower() or token in msg_lower
            for token in ("parse", "analysis", "syntax", "sqlstate", "mismatched input")
        )
        traceback_clean = _suppress_jvm_stacktrace(traceback.format_exc())
        raise HTTPException(
            status_code=400 if is_user_sql_error else 500,
            detail={
                "message": "Spark SQL export failed",
                "error_type": exc_name,
                "spark_message": spark_message_clean,
                "database": body.database,
                "sql": body.sql,
                "traceback": traceback_clean,
            },
        )


@router.post("/query/export/jobs")
async def create_query_export_job(body: QueryExportRequest):
    _check_sql_allowed(body.sql)

    job_id = uuid.uuid4().hex
    now = _utc_now_iso()
    _EXPORT_JOBS[job_id] = {
        "id": job_id,
        "status": "queued",
        "phase": "queued",
        "format": body.format,
        "rows_exported": 0,
        "max_rows": body.max_rows,
        "database": body.database,
        "sql": body.sql,
        "file_path": None,
        "error": None,
        "cancel_requested": False,
        "created_at": now,
        "started_at": None,
        "completed_at": None,
        "updated_at": now,
    }
    asyncio.create_task(_run_query_export_job(job_id))
    return _serialize_export_job(_EXPORT_JOBS[job_id])


@router.get("/query/export/jobs/{job_id}")
async def get_query_export_job(job_id: str):
    job = _EXPORT_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Export job not found")
    return _serialize_export_job(job)


@router.post("/query/export/jobs/{job_id}/cancel")
async def cancel_query_export_job(job_id: str):
    job = _EXPORT_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Export job not found")
    if job["status"] in {"completed", "failed", "canceled"}:
        return _serialize_export_job(job)
    job["cancel_requested"] = True
    job["status"] = "canceling"
    job["phase"] = "canceling"
    job["updated_at"] = _utc_now_iso()
    return _serialize_export_job(job)


@router.get("/query/export/jobs/{job_id}/download")
async def download_query_export_job(job_id: str):
    job = _EXPORT_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Export job not found")
    if job["status"] != "completed" or not job.get("file_path"):
        raise HTTPException(status_code=409, detail="Export job is not ready for download")
    path = Path(job["file_path"])
    if not path.exists():
        raise HTTPException(status_code=410, detail="Export file no longer exists")
    fmt = job.get("format", "csv")
    if fmt == "xlsx":
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"query_server_export_{job_id[:8]}.xlsx"
    else:
        media_type = "text/csv"
        filename = f"query_server_export_{job_id[:8]}.csv"
    return FileResponse(path, media_type=media_type, filename=filename)


# ─────────────────────────────────────────────────────────────────────────────
# Spark Catalog — databases and tables
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/catalog/databases")
async def list_catalog_databases():
    """List all Spark databases."""
    try:
        return await spark_service.list_databases()
    except Exception as exc:
        logger.exception("Error listing Spark databases")
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/catalog/databases/{database}/introspect")
async def introspect_catalog_database(database: str):
    """Return table and column metadata for a Spark database."""
    try:
        tables = await spark_service.list_catalog_tables()
        database_tables = sorted(
            [t["name"] for t in tables if t.get("database") == database and not t.get("is_temporary")]
        )

        result: list[dict] = []
        for table_name in database_tables:
            describe = await spark_service.execute_query(
                f"DESCRIBE `{database}`.`{table_name}`",
                limit=1000,
                offset=0,
                database=database,
            )
            columns: list[dict] = []
            for row in describe.get("rows", []):
                col_name = str(row[0] or "").strip() if len(row) > 0 else ""
                data_type = str(row[1] or "").strip() if len(row) > 1 else ""
                if not col_name or col_name.startswith("#"):
                    continue
                columns.append({"name": col_name, "type": data_type})
            result.append({"name": table_name, "columns": columns})

        return {"database": database, "tables": result}
    except Exception as exc:
        logger.exception("Error introspecting Spark database %s", database)
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/catalog/tables")
async def list_catalog_tables():
    """List all Spark catalog tables and views across all databases."""
    try:
        return await spark_service.list_catalog_tables()
    except Exception as exc:
        logger.exception("Error listing catalog tables")
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/catalog/reconnect")
async def catalog_reconnect():
    """Drop the current Spark session and establish a fresh connection."""
    reset_spark_session()
    try:
        def _connect():
            spark = _get_spark()
            return spark.range(1).collect()
        await asyncio.to_thread(_connect)
        return {"status": "connected"}
    except Exception as exc:
        logger.exception("Spark reconnect failed")
        raise HTTPException(status_code=503, detail=str(exc))


@router.post("/catalog/disconnect")
async def catalog_disconnect():
    """Drop the current Spark Connect session."""
    reset_spark_session()
    return {"status": "disconnected"}


@router.delete("/catalog/views/{view_name}")
async def drop_temp_view(view_name: str):
    """Drop a temporary view by name."""
    try:
        await spark_service.execute_query(f"DROP VIEW IF EXISTS `{view_name}`")
        return {"status": "dropped", "view": view_name}
    except Exception as exc:
        logger.exception("Error dropping view %s", view_name)
        raise HTTPException(status_code=500, detail=str(exc))


@router.delete("/catalog/views")
async def drop_all_temp_views():
    """Drop all temporary views in the current Spark session."""
    try:
        tables = await spark_service.list_catalog_tables()
        temp_views = [t["name"] for t in tables if t.get("is_temporary")]
        for name in temp_views:
            await spark_service.execute_query(f"DROP VIEW IF EXISTS `{name}`")
        return {"status": "dropped", "count": len(temp_views), "views": temp_views}
    except Exception as exc:
        logger.exception("Error dropping all temp views")
        raise HTTPException(status_code=500, detail=str(exc))
